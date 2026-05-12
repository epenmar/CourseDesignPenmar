"""Lightweight report export builders for existing v2 session tables."""

from __future__ import annotations

import csv
import io
import json
import re
from collections import Counter
from datetime import datetime, timezone
from typing import Any


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _csv_bytes(rows: list[dict[str, Any]], headers: list[str]) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in headers})
    return output.getvalue().encode("utf-8")


def _json_bytes(payload: dict[str, Any]) -> bytes:
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def _safe(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def _fetch_all(query, *, page_size: int = 1000) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    offset = 0
    while True:
        result = query.range(offset, offset + page_size - 1).execute()
        batch = result.data or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return rows


def _content_status(row: dict[str, Any]) -> str:
    metadata = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
    linked_from = metadata.get("linked_from") if isinstance(metadata.get("linked_from"), list) else []
    if row.get("duplicate_group_key"):
        return "duplicate"
    if linked_from:
        return "referenced"
    if row.get("is_orphaned") and row.get("content_type") != "module":
        return "orphaned"
    if row.get("published") is False:
        return "unpublished"
    return "in_module"


def _decision_map(supabase, *, session_id: str, user_id: str, content_item_ids: list[str]) -> dict[str, dict[str, Any]]:
    if not content_item_ids:
        return {}
    result = supabase.table("content_inventory_decisions").select(
        "content_item_id, action, reason, applied_to_canvas, applied_at, updated_at"
    ).eq("session_id", session_id).eq("user_id", user_id).in_(
        "content_item_id", content_item_ids
    ).execute()
    return {
        row["content_item_id"]: row
        for row in result.data or []
        if row.get("content_item_id")
    }


def _xlsx_bytes(wb) -> bytes:
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _workbook_helpers():
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    maroon = "8C1D40"
    light_green = "E2EFDA"
    light_red = "FCE4EC"
    light_yellow = "FFF2CC"
    light_blue = "D6E4F0"
    light_gray = "F2F2F2"
    banner_fill = PatternFill("solid", fgColor="FFF2CC")
    banner_font = Font(bold=True, color="5C3B00", size=11, name="Calibri")
    section_font = Font(bold=True, color=maroon, size=12, name="Calibri")
    header_fill = PatternFill("solid", fgColor=maroon)
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    body_font = Font(size=10, name="Calibri")
    link_font = Font(color="0563C1", underline="single", size=10, name="Calibri")
    body_align = Alignment(vertical="top", wrap_text=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    status_fills = {
        "in_module": PatternFill("solid", fgColor=light_green),
        "referenced": PatternFill("solid", fgColor=light_green),
        "linked": PatternFill("solid", fgColor=light_blue),
        "unpublished": PatternFill("solid", fgColor=light_gray),
        "orphaned": PatternFill("solid", fgColor=light_red),
        "unreferenced": PatternFill("solid", fgColor=light_red),
        "not_in_module": PatternFill("solid", fgColor=light_red),
        "duplicate": PatternFill("solid", fgColor=light_yellow),
        "warning": PatternFill("solid", fgColor=light_yellow),
        "critical": PatternFill("solid", fgColor=light_red),
        "info": PatternFill("solid", fgColor=light_gray),
    }
    decision_fills = {
        "keep": PatternFill("solid", fgColor=light_green),
        "delete": PatternFill("solid", fgColor=light_red),
        "defer": PatternFill("solid", fgColor=light_yellow),
    }
    return {
        "Workbook": Workbook,
        "DataValidation": DataValidation,
        "OpenpyxlImage": OpenpyxlImage,
        "get_column_letter": get_column_letter,
        "banner_fill": banner_fill,
        "banner_font": banner_font,
        "section_font": section_font,
        "header_fill": header_fill,
        "header_font": header_font,
        "body_font": body_font,
        "link_font": link_font,
        "body_align": body_align,
        "header_align": header_align,
        "thin_border": thin_border,
        "status_fills": status_fills,
        "decision_fills": decision_fills,
    }


def _style_header_row(ws, headers: list[str], helpers: dict[str, Any], row: int = 1) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = helpers["header_fill"]
        cell.font = helpers["header_font"]
        cell.alignment = helpers["header_align"]
        cell.border = helpers["thin_border"]
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{row}:{helpers['get_column_letter'](len(headers))}{row}"


def _write_cell(ws, row: int, col: int, value: Any, helpers: dict[str, Any], *, hyperlink: str | None = None, fill_key: str | None = None) -> None:
    cell_value = value if isinstance(value, (int, float)) and not isinstance(value, bool) else _safe(value)
    cell = ws.cell(row=row, column=col, value=cell_value)
    cell.font = helpers["link_font"] if hyperlink else helpers["body_font"]
    cell.alignment = helpers["body_align"]
    cell.border = helpers["thin_border"]
    if hyperlink:
        cell.hyperlink = hyperlink
    if fill_key:
        fill = helpers["status_fills"].get(fill_key) or helpers["decision_fills"].get(fill_key)
        if fill:
            cell.fill = fill


def _auto_col_widths(ws, helpers: dict[str, Any], *, min_w: int = 10, max_w: int = 55) -> None:
    for col_cells in ws.columns:
        col_letter = helpers["get_column_letter"](col_cells[0].column)
        max_len = max(min(len(str(cell.value or "")), max_w) for cell in col_cells)
        ws.column_dimensions[col_letter].width = max(min_w, max_len + 2)


def _decision_label(action: Any) -> str:
    if action == "delete":
        return "Remove"
    if action == "defer":
        return "Defer"
    if action == "keep":
        return "Keep"
    return ""


def _content_inventory_rows(supabase, *, session_id: str, user_id: str) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    items = _fetch_all(
        supabase.table("course_content_items").select(
            "id, canvas_id, content_type, title, canvas_url, published, module_name, is_orphaned, duplicate_group_key, metadata, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).order("content_type").order("title")
    )
    decisions = _decision_map(
        supabase,
        session_id=session_id,
        user_id=user_id,
        content_item_ids=[row["id"] for row in items if row.get("id")],
    )
    return items, decisions


def _health_findings(supabase, *, session_id: str, user_id: str) -> list[dict[str, Any]]:
    run_result = supabase.table("health_runs").select("id").eq("session_id", session_id).eq(
        "user_id", user_id
    ).order("created_at", desc=True).limit(1).execute()
    if not run_result.data:
        return []
    return _fetch_all(
        supabase.table("health_findings").select(
            "finding_type, finding_code, severity, description, content_item_id, context, created_at"
        ).eq("session_id", session_id).eq("health_run_id", run_result.data[0]["id"]).order("severity")
    )


def _add_decision_validation(ws, helpers: dict[str, Any], cell_range: str) -> None:
    dv = helpers["DataValidation"](
        type="list",
        formula1='"Keep,Remove"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Choose Keep or Remove"
    dv.errorTitle = "Invalid value"
    dv.prompt = "Keep or Remove"
    dv.promptTitle = "Review Decision"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _add_banner(ws, text: str, helpers: dict[str, Any], *, columns: int, row: int = 1) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = helpers["banner_fill"]
    cell.font = helpers["banner_font"]
    cell.alignment = helpers["body_align"]
    ws.row_dimensions[row].height = 34
    return row + 1


def _metadata(row: dict[str, Any] | None) -> dict[str, Any]:
    value = (row or {}).get("metadata")
    return value if isinstance(value, dict) else {}


def _linked_from(row: dict[str, Any] | None) -> list[str]:
    value = _metadata(row).get("linked_from")
    return [str(item) for item in value] if isinstance(value, list) else []


def _course_context(supabase, *, session_id: str, user_id: str) -> dict[str, Any]:
    session_result = supabase.table("sessions").select(
        "id, name, source_course_id, updated_at"
    ).eq("id", session_id).eq("user_id", user_id).limit(1).execute()
    session = session_result.data[0] if session_result.data else {}
    course: dict[str, Any] = {}
    if session.get("source_course_id"):
        course_result = supabase.table("courses").select(
            "course_name, canvas_base_url, canvas_course_id"
        ).eq("id", session["source_course_id"]).eq("user_id", user_id).limit(1).execute()
        course = course_result.data[0] if course_result.data else {}
    course_url = ""
    if course.get("canvas_base_url") and course.get("canvas_course_id"):
        course_url = f"{str(course['canvas_base_url']).rstrip('/')}/courses/{course['canvas_course_id']}"
    return {
        "name": course.get("course_name") or session.get("name") or "Course",
        "url": course_url,
        "audit_date": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


def _is_graded(row: dict[str, Any]) -> bool:
    metadata = _metadata(row)
    if row.get("content_type") in {"assignment", "quiz"}:
        return True
    if row.get("content_type") == "discussion":
        return bool(metadata.get("is_discussion_assignment") or metadata.get("assignment_id"))
    points = metadata.get("points_possible") or metadata.get("assignment_shell_points_possible")
    try:
        return float(points or 0) > 0
    except (TypeError, ValueError):
        return False


def _item_location(row: dict[str, Any] | None) -> str:
    if not row:
        return ""
    metadata = _metadata(row)
    linked = _linked_from(row)
    return (
        row.get("module_name")
        or metadata.get("parent_quiz_title")
        or metadata.get("folder_path")
        or metadata.get("folder_name")
        or ("; ".join(linked[:5]) if linked else "")
        or ""
    )


def _item_key(row: dict[str, Any]) -> str:
    return f"{row.get('content_type') or ''}-{row.get('canvas_id') or ''}"


def _context_text(context: Any) -> str:
    if not isinstance(context, dict):
        return ""
    return "; ".join(f"{key}: {value}" for key, value in context.items() if value not in (None, ""))


def _finding_location(finding: dict[str, Any]) -> str:
    context = finding.get("context") if isinstance(finding.get("context"), dict) else {}
    for key, label in [
        ("image_index", "Image"),
        ("link_index", "Link"),
        ("heading_index", "Heading"),
        ("table_index", "Table"),
    ]:
        if context.get(key):
            return f"{label} {context[key]}"
    if context.get("tag"):
        return str(context["tag"]).upper()
    return ""


def _finding_snippet(finding: dict[str, Any]) -> str:
    context = finding.get("context") if isinstance(finding.get("context"), dict) else {}
    if context.get("src"):
        return f'<img src="{context["src"]}">'
    if context.get("href"):
        text = context.get("text") or ""
        return f'<a href="{context["href"]}">{text}</a>'
    if context.get("text"):
        return str(context["text"])
    return _context_text(context)


def _image_issue_type(image: dict[str, Any]) -> str:
    if image.get("is_broken"):
        return "Broken image"
    if not image.get("is_decorative") and not str(image.get("existing_alt_text") or "").strip():
        return "Missing alt text"
    if image.get("review_action") == "delete":
        return "Marked for removal"
    return ""


def _image_display_name(image: dict[str, Any], file_by_canvas_id: dict[str, dict[str, Any]]) -> str:
    if image.get("canvas_file_id"):
        file_item = file_by_canvas_id.get(str(image["canvas_file_id"]))
        if file_item:
            metadata = _metadata(file_item)
            return str(metadata.get("filename") or file_item.get("title") or image["canvas_file_id"])
    url = str(image.get("canvas_url") or "")
    if not url:
        return str(image.get("canvas_file_id") or "Image")
    parts = [part for part in url.split("?")[0].rstrip("/").split("/") if part]
    tail = parts[-1] if parts else ""
    if tail.lower() in {"preview", "download"} or tail.isdigit():
        return str(image.get("canvas_file_id") or "Image")
    return tail or str(image.get("canvas_file_id") or "Image")


def _limited_text(value: Any, limit: int = 500) -> str:
    text = _safe(value)
    return text if len(text) <= limit else f"{text[:limit - 1]}..."


def _embed_thumbnail(ws, *, row: int, col: int, image: dict[str, Any], helpers: dict[str, Any]) -> None:
    if row > 101:
        return
    thumb_key = image.get("r2_thumb_key")
    if not thumb_key:
        return
    try:
        from r2_storage import download_bytes, is_r2_configured

        if not is_r2_configured():
            return
        data, _ = download_bytes(str(thumb_key))
        drawing = helpers["OpenpyxlImage"](io.BytesIO(data))
        drawing.width = min(drawing.width or 120, 140)
        drawing.height = min(drawing.height or 90, 100)
        ws.add_image(drawing, ws.cell(row=row, column=col).coordinate)
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 20, 78)
    except Exception:
        return


def _sheet_summary(
    wb,
    helpers: dict[str, Any],
    *,
    course: dict[str, Any],
    items: list[dict[str, Any]],
    images: list[dict[str, Any]],
    findings: list[dict[str, Any]],
) -> None:
    summary_ws = wb.create_sheet("Summary", 0)
    counts: dict[str, int] = {}
    for item in items:
        content_type = str(item.get("content_type") or "unknown")
        counts[content_type] = counts.get(content_type, 0) + 1
    finding_counts = Counter(str(row.get("finding_code") or row.get("finding_type") or "unknown") for row in findings)
    image_alt_issue_count = (
        finding_counts.get("missing_image_alt", 0)
        + finding_counts.get("generic_image_alt", 0)
        + finding_counts.get("filename_image_alt", 0)
        + finding_counts.get("image_alt_too_long", 0)
    )
    orphan_counts = Counter(str(item.get("content_type") or "unknown") for item in items if item.get("is_orphaned"))
    summary_rows = [
        ("Course Name", course.get("name") or ""),
        ("Course URL", course.get("url") or ""),
        ("Audit Date", course.get("audit_date") or ""),
        ("", ""),
        ("CONTENT COUNTS", ""),
        ("Content Items", len(items)),
        ("Pages", counts.get("page", 0)),
        ("Assignments", counts.get("assignment", 0)),
        ("Discussions", counts.get("discussion", 0)),
        ("Quizzes", counts.get("quiz", 0)),
        ("Files", counts.get("file", 0)),
        ("Modules", counts.get("module", 0)),
        ("", ""),
        ("ACCESSIBILITY ISSUES", ""),
        ("Images", len(images)),
        ("Image Alt Text Issues", image_alt_issue_count),
        ("WCAG Issues", sum(1 for row in findings if row.get("finding_type") == "wcag")),
        ("Link Text Issues", finding_counts.get("empty_link_text", 0) + finding_counts.get("generic_link_text", 0)),
        ("Latest Health Findings", len(findings)),
        ("", ""),
        ("ORPHANED CONTENT", ""),
        ("Orphaned Pages", orphan_counts.get("page", 0)),
        ("Orphaned Assignments", orphan_counts.get("assignment", 0)),
        ("Orphaned Discussions", orphan_counts.get("discussion", 0)),
        ("Orphaned Quizzes", orphan_counts.get("quiz", 0)),
        ("Unreferenced Files", orphan_counts.get("file", 0)),
        ("Total Orphaned", sum(orphan_counts.values())),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, 1):
        _write_cell(summary_ws, row_idx, 1, label, helpers)
        _write_cell(summary_ws, row_idx, 2, value, helpers)
        if label and label == label.upper():
            summary_ws.cell(row=row_idx, column=1).font = helpers["section_font"]
        if label == "Course URL" and value:
            summary_ws.cell(row=row_idx, column=2).hyperlink = str(value)
            summary_ws.cell(row=row_idx, column=2).font = helpers["link_font"]
    _auto_col_widths(summary_ws, helpers)


def _sheet_content_inventory(
    wb,
    helpers: dict[str, Any],
    *,
    items: list[dict[str, Any]],
    decisions: dict[str, dict[str, Any]],
) -> None:
    content_ws = wb.create_sheet("Content Inventory")
    content_headers = [
        "Type", "Title", "Module", "Status", "Faculty Decision", "Published",
        "Graded", "Linked From", "Canvas Link", "Item Key (do not edit)",
    ]
    header_row = _add_banner(
        content_ws,
        "FACULTY REVIEW — The 'Faculty Decision' column is pre-filled with the tool's recommendation based on each item's status. Click any cell to change it to Keep or Remove. You only need to override the items you disagree with — when in doubt, keep it.",
        helpers,
        columns=len(content_headers),
    )
    _style_header_row(content_ws, content_headers, helpers, row=header_row)
    first_data_row = header_row + 1
    for offset, item in enumerate(items):
        row_idx = first_data_row + offset
        linked_from = _linked_from(item)
        status = _content_status(item)
        decision = decisions.get(item.get("id")) or {}
        _write_cell(content_ws, row_idx, 1, str(item.get("content_type") or "").title(), helpers)
        _write_cell(content_ws, row_idx, 2, item.get("title") or "Untitled", helpers)
        _write_cell(content_ws, row_idx, 3, _item_location(item), helpers)
        _write_cell(content_ws, row_idx, 4, status.replace("_", " ").title(), helpers, fill_key=status)
        _write_cell(content_ws, row_idx, 5, _decision_label(decision.get("action")), helpers, fill_key=decision.get("action"))
        _write_cell(content_ws, row_idx, 6, item.get("published"), helpers)
        _write_cell(content_ws, row_idx, 7, _is_graded(item), helpers)
        _write_cell(content_ws, row_idx, 8, ", ".join(linked_from), helpers)
        _write_cell(content_ws, row_idx, 9, "Open in Canvas" if item.get("canvas_url") else "", helpers, hyperlink=item.get("canvas_url"))
        _write_cell(content_ws, row_idx, 10, _item_key(item), helpers)
    if items:
        _add_decision_validation(content_ws, helpers, f"E{first_data_row}:E{first_data_row + len(items) - 1}")
    _auto_col_widths(content_ws, helpers)
    content_ws.column_dimensions[helpers["get_column_letter"](len(content_headers))].hidden = True


def _sheet_wcag_issues(
    wb,
    helpers: dict[str, Any],
    *,
    findings: list[dict[str, Any]],
    item_by_id: dict[str, dict[str, Any]],
) -> None:
    wcag_ws = wb.create_sheet("WCAG Issues")
    wcag_headers = ["Severity", "Content Type", "Page Title", "Location", "Issue", "HTML Snippet", "Canvas Link"]
    _style_header_row(wcag_ws, wcag_headers, helpers)
    wcag_findings = [finding for finding in findings if finding.get("finding_type") == "wcag"]
    for row_idx, finding in enumerate(wcag_findings, 2):
        item = item_by_id.get(str(finding.get("content_item_id") or "")) or {}
        severity = str(finding.get("severity") or "")
        _write_cell(wcag_ws, row_idx, 1, severity.title(), helpers, fill_key=severity)
        _write_cell(wcag_ws, row_idx, 2, str(item.get("content_type") or "").title(), helpers)
        _write_cell(wcag_ws, row_idx, 3, item.get("title") or "", helpers)
        _write_cell(wcag_ws, row_idx, 4, _finding_location(finding) or _item_location(item), helpers)
        _write_cell(wcag_ws, row_idx, 5, finding.get("description") or finding.get("finding_code"), helpers)
        _write_cell(wcag_ws, row_idx, 6, _limited_text(_finding_snippet(finding), 500), helpers)
        _write_cell(wcag_ws, row_idx, 7, "Open in Canvas" if item.get("canvas_url") else "", helpers, hyperlink=item.get("canvas_url"))
    _auto_col_widths(wcag_ws, helpers)


def _sheet_image_alt_text(
    wb,
    helpers: dict[str, Any],
    *,
    images: list[dict[str, Any]],
    item_by_id: dict[str, dict[str, Any]],
    file_by_canvas_id: dict[str, dict[str, Any]],
) -> None:
    image_ws = wb.create_sheet("Image Alt Text")
    image_headers = ["Thumbnail", "Filename", "Current Alt Text", "Issue Type", "Page Context", "Location", "Decorative", "Canvas Link"]
    _style_header_row(image_ws, image_headers, helpers)
    for row_idx, image in enumerate(images, 2):
        item = item_by_id.get(str(image.get("content_item_id") or "")) or {}
        _write_cell(image_ws, row_idx, 1, "", helpers)
        _embed_thumbnail(image_ws, row=row_idx, col=1, image=image, helpers=helpers)
        _write_cell(image_ws, row_idx, 2, _image_display_name(image, file_by_canvas_id), helpers, hyperlink=image.get("canvas_url"))
        _write_cell(image_ws, row_idx, 3, image.get("existing_alt_text"), helpers)
        _write_cell(image_ws, row_idx, 4, _image_issue_type(image), helpers)
        _write_cell(image_ws, row_idx, 5, item.get("title") or "", helpers)
        _write_cell(image_ws, row_idx, 6, _item_location(item), helpers)
        _write_cell(image_ws, row_idx, 7, image.get("is_decorative"), helpers)
        _write_cell(image_ws, row_idx, 8, "Open in Canvas" if item.get("canvas_url") else "", helpers, hyperlink=item.get("canvas_url"))
    _auto_col_widths(image_ws, helpers)
    image_ws.column_dimensions["A"].width = 20


def _sheet_link_issues(
    wb,
    helpers: dict[str, Any],
    *,
    findings: list[dict[str, Any]],
    item_by_id: dict[str, dict[str, Any]],
) -> None:
    link_ws = wb.create_sheet("Link Issues")
    link_headers = ["Link Text", "URL", "Issue Type", "Page Title", "Location", "Context", "Canvas Link"]
    _style_header_row(link_ws, link_headers, helpers)
    link_findings = [
        finding for finding in findings
        if "link" in str(finding.get("finding_code") or "").casefold()
    ]
    for row_idx, finding in enumerate(link_findings, 2):
        item = item_by_id.get(str(finding.get("content_item_id") or "")) or {}
        context = finding.get("context") if isinstance(finding.get("context"), dict) else {}
        _write_cell(link_ws, row_idx, 1, context.get("text") or "", helpers)
        _write_cell(link_ws, row_idx, 2, context.get("href") or "", helpers, hyperlink=context.get("href"))
        _write_cell(link_ws, row_idx, 3, str(finding.get("finding_code") or "").replace("_", " ").title(), helpers)
        _write_cell(link_ws, row_idx, 4, item.get("title") or "", helpers)
        _write_cell(link_ws, row_idx, 5, _finding_location(finding) or _item_location(item), helpers)
        _write_cell(link_ws, row_idx, 6, _limited_text(finding.get("description") or _context_text(context), 500), helpers)
        _write_cell(link_ws, row_idx, 7, "Open in Canvas" if item.get("canvas_url") else "", helpers, hyperlink=item.get("canvas_url"))
    _auto_col_widths(link_ws, helpers)


def _sheet_health_summary(
    wb,
    helpers: dict[str, Any],
    *,
    course: dict[str, Any],
    run: dict[str, Any] | None,
    items: list[dict[str, Any]],
    images: list[dict[str, Any]],
    findings: list[dict[str, Any]],
) -> None:
    summary_ws = wb.create_sheet("Summary", 0)
    summary = run.get("summary") if isinstance(run, dict) and isinstance(run.get("summary"), dict) else {}
    counts = Counter(str(item.get("content_type") or "unknown") for item in items)
    finding_counts = Counter(str(row.get("finding_code") or row.get("finding_type") or "unknown") for row in findings)
    severity_counts = Counter(str(row.get("severity") or "info") for row in findings)
    image_alt_issue_count = (
        finding_counts.get("missing_image_alt", 0)
        + finding_counts.get("generic_image_alt", 0)
        + finding_counts.get("filename_image_alt", 0)
        + finding_counts.get("image_alt_too_long", 0)
    )
    rows = [
        ("Course Name", course.get("name") or ""),
        ("Course URL", course.get("url") or ""),
        ("Audit Date", course.get("audit_date") or ""),
        ("", ""),
        ("HEALTH RUN", ""),
        ("Status", run.get("status") if run else "Not available"),
        ("Items Scanned", run.get("items_scanned") if run else 0),
        ("Duration (ms)", run.get("duration_ms") if run else ""),
        ("Started", run.get("created_at") if run else ""),
        ("Finished", run.get("finished_at") if run else ""),
        ("", ""),
        ("FINDINGS", ""),
        ("Total Findings", len(findings)),
        ("Critical", severity_counts.get("critical", 0)),
        ("Warnings", severity_counts.get("warning", 0)),
        ("Info", severity_counts.get("info", 0)),
        ("Image Alt Text Issues", image_alt_issue_count),
        ("Link Text Issues", finding_counts.get("empty_link_text", 0) + finding_counts.get("generic_link_text", 0)),
        ("Heading Issues", finding_counts.get("empty_heading", 0) + finding_counts.get("skipped_heading_level", 0)),
        ("Table Issues", finding_counts.get("table_missing_header", 0)),
        ("Duplicate Content", finding_counts.get("duplicate_content", 0)),
        ("Orphaned Content", finding_counts.get("orphaned_content", 0)),
        ("Unpublished Content", finding_counts.get("unpublished_content", 0)),
        ("", ""),
        ("CONTENT COUNTS", ""),
        ("Content Items", len(items)),
        ("Pages", counts.get("page", 0)),
        ("Assignments", counts.get("assignment", 0)),
        ("Discussions", counts.get("discussion", 0)),
        ("Quizzes", counts.get("quiz", 0)),
        ("Files", counts.get("file", 0)),
        ("Images", len(images)),
    ]
    for key, value in sorted(summary.items()):
        if key in {"total_findings", "by_code", "by_severity"}:
            continue
        rows.append((str(key).replace("_", " ").title(), value))

    for row_idx, (label, value) in enumerate(rows, 1):
        _write_cell(summary_ws, row_idx, 1, label, helpers)
        _write_cell(summary_ws, row_idx, 2, value, helpers)
        if label and label == label.upper():
            summary_ws.cell(row=row_idx, column=1).font = helpers["section_font"]
        if label == "Course URL" and value:
            summary_ws.cell(row=row_idx, column=2).hyperlink = str(value)
            summary_ws.cell(row=row_idx, column=2).font = helpers["link_font"]
    _auto_col_widths(summary_ws, helpers)


def _sheet_health_image_issues(
    wb,
    helpers: dict[str, Any],
    *,
    images: list[dict[str, Any]],
    item_by_id: dict[str, dict[str, Any]],
    file_by_canvas_id: dict[str, dict[str, Any]],
) -> None:
    image_ws = wb.create_sheet("Image Issues")
    headers = [
        "Thumbnail", "Filename", "Issue Type", "Current Alt Text", "Reviewed Alt Text",
        "Long Description", "Page Context", "Location", "Canvas Link", "Image URL",
    ]
    _style_header_row(image_ws, headers, helpers)
    issue_images = [image for image in images if _image_issue_type(image)]
    for row_idx, image in enumerate(issue_images, 2):
        item = item_by_id.get(str(image.get("content_item_id") or "")) or {}
        _write_cell(image_ws, row_idx, 1, "", helpers)
        _embed_thumbnail(image_ws, row=row_idx, col=1, image=image, helpers=helpers)
        _write_cell(image_ws, row_idx, 2, _image_display_name(image, file_by_canvas_id), helpers, hyperlink=image.get("canvas_url"))
        _write_cell(image_ws, row_idx, 3, _image_issue_type(image), helpers, fill_key="warning")
        _write_cell(image_ws, row_idx, 4, image.get("existing_alt_text"), helpers)
        _write_cell(image_ws, row_idx, 5, image.get("edited_alt_text"), helpers)
        _write_cell(image_ws, row_idx, 6, image.get("long_description"), helpers)
        _write_cell(image_ws, row_idx, 7, item.get("title") or "", helpers)
        _write_cell(image_ws, row_idx, 8, _item_location(item), helpers)
        _write_cell(image_ws, row_idx, 9, "Open in Canvas" if item.get("canvas_url") else "", helpers, hyperlink=item.get("canvas_url"))
        _write_cell(image_ws, row_idx, 10, image.get("canvas_url"), helpers, hyperlink=image.get("canvas_url"))
    _auto_col_widths(image_ws, helpers)
    image_ws.column_dimensions["A"].width = 20


def _sheet_health_files_documents(
    wb,
    helpers: dict[str, Any],
    *,
    file_items: list[dict[str, Any]],
    documents: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Files and Documents")
    headers = ["Type", "Name", "Status", "Location", "Size (KB)", "MIME Type", "Pages", "Canvas Link", "Updated At"]
    _style_header_row(ws, headers, helpers)
    row_idx = 2
    for item in file_items:
        metadata = _metadata(item)
        size_bytes = metadata.get("size")
        try:
            size_kb = round(float(size_bytes or 0) / 1024, 1) if size_bytes is not None else ""
        except (TypeError, ValueError):
            size_kb = ""
        status = _content_status(item)
        _write_cell(ws, row_idx, 1, "Canvas File", helpers)
        _write_cell(ws, row_idx, 2, metadata.get("filename") or item.get("title") or "Untitled file", helpers)
        _write_cell(ws, row_idx, 3, status.replace("_", " ").title(), helpers, fill_key=status)
        _write_cell(ws, row_idx, 4, metadata.get("folder_path") or metadata.get("folder_name") or _item_location(item), helpers)
        _write_cell(ws, row_idx, 5, size_kb, helpers)
        _write_cell(ws, row_idx, 6, metadata.get("content_type") or "", helpers)
        _write_cell(ws, row_idx, 7, "", helpers)
        _write_cell(ws, row_idx, 8, "Open in Canvas" if item.get("canvas_url") else "", helpers, hyperlink=item.get("canvas_url"))
        _write_cell(ws, row_idx, 9, item.get("updated_at"), helpers)
        row_idx += 1

    for document in documents:
        tag_data = document.get("tag_data") if isinstance(document.get("tag_data"), dict) else {}
        size_bytes = tag_data.get("size")
        try:
            size_kb = round(float(size_bytes or 0) / 1024, 1) if size_bytes is not None else ""
        except (TypeError, ValueError):
            size_kb = ""
        _write_cell(ws, row_idx, 1, "Document", helpers)
        _write_cell(ws, row_idx, 2, document.get("filename") or "Untitled document", helpers)
        _write_cell(ws, row_idx, 3, document.get("status") or "", helpers)
        _write_cell(ws, row_idx, 4, tag_data.get("source") or "", helpers)
        _write_cell(ws, row_idx, 5, size_kb, helpers)
        _write_cell(ws, row_idx, 6, tag_data.get("mime_type") or "", helpers)
        _write_cell(ws, row_idx, 7, document.get("page_count") or "", helpers)
        _write_cell(ws, row_idx, 8, tag_data.get("canvas_url") or "", helpers, hyperlink=tag_data.get("canvas_url"))
        _write_cell(ws, row_idx, 9, document.get("updated_at") or document.get("created_at"), helpers)
        row_idx += 1
    _auto_col_widths(ws, helpers)


def _sheet_health_inventory_findings(
    wb,
    helpers: dict[str, Any],
    *,
    findings: list[dict[str, Any]],
    item_by_id: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Inventory Findings")
    headers = ["Issue Type", "Severity", "Content Type", "Title", "Status", "Location", "Description", "Canvas Link"]
    _style_header_row(ws, headers, helpers)
    rows = [finding for finding in findings if finding.get("finding_type") == "inventory"]
    for row_idx, finding in enumerate(rows, 2):
        item = item_by_id.get(str(finding.get("content_item_id") or "")) or {}
        severity = str(finding.get("severity") or "info")
        status = _content_status(item) if item else ""
        _write_cell(ws, row_idx, 1, str(finding.get("finding_code") or "").replace("_", " ").title(), helpers)
        _write_cell(ws, row_idx, 2, severity.title(), helpers, fill_key=severity)
        _write_cell(ws, row_idx, 3, str(item.get("content_type") or "").title(), helpers)
        _write_cell(ws, row_idx, 4, item.get("title") or "", helpers)
        _write_cell(ws, row_idx, 5, status.replace("_", " ").title(), helpers, fill_key=status)
        _write_cell(ws, row_idx, 6, _item_location(item), helpers)
        _write_cell(ws, row_idx, 7, finding.get("description") or "", helpers)
        _write_cell(ws, row_idx, 8, "Open in Canvas" if item.get("canvas_url") else "", helpers, hyperlink=item.get("canvas_url"))
    _auto_col_widths(ws, helpers)


def _sheet_files(
    wb,
    helpers: dict[str, Any],
    *,
    items: list[dict[str, Any]],
    decisions: dict[str, dict[str, Any]],
) -> None:
    file_ws = wb.create_sheet("Files")
    file_items = [item for item in items if item.get("content_type") == "file"]
    file_headers = ["Filename", "Folder / Location", "Status", "Faculty Decision", "Size (KB)", "MIME Type", "Canvas Link", "Item Key (do not edit)"]
    _style_header_row(file_ws, file_headers, helpers)
    for row_idx, item in enumerate(file_items, 2):
        metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
        status = _content_status(item)
        decision = decisions.get(item.get("id")) or {}
        size_bytes = metadata.get("size")
        try:
            size_kb = round(float(size_bytes or 0) / 1024, 1) if size_bytes is not None else ""
        except (TypeError, ValueError):
            size_kb = ""
        _write_cell(file_ws, row_idx, 1, metadata.get("filename") or item.get("title") or "Untitled file", helpers)
        _write_cell(file_ws, row_idx, 2, metadata.get("folder_path") or metadata.get("folder_name") or item.get("module_name") or "", helpers)
        _write_cell(file_ws, row_idx, 3, status.replace("_", " ").title(), helpers, fill_key=status)
        _write_cell(file_ws, row_idx, 4, _decision_label(decision.get("action")), helpers, fill_key=decision.get("action"))
        _write_cell(file_ws, row_idx, 5, size_kb, helpers)
        _write_cell(file_ws, row_idx, 6, metadata.get("content_type") or "", helpers)
        _write_cell(file_ws, row_idx, 7, "Open in Canvas" if item.get("canvas_url") else "", helpers, hyperlink=item.get("canvas_url"))
        _write_cell(file_ws, row_idx, 8, f"file-{item.get('canvas_id')}", helpers)
    if file_items:
        _add_decision_validation(file_ws, helpers, f"D2:D{len(file_items) + 1}")
    _auto_col_widths(file_ws, helpers)
    file_ws.column_dimensions[helpers["get_column_letter"](len(file_headers))].hidden = True


def _format_answers(answers: Any) -> str:
    if not isinstance(answers, list):
        return ""
    labels: list[str] = []
    for idx, answer in enumerate(answers, 1):
        if not isinstance(answer, dict):
            continue
        raw_text = (
            answer.get("text")
            or answer.get("html")
            or answer.get("answer_text")
            or answer.get("left")
            or answer.get("right")
            or answer.get("exact")
            or ""
        )
        text = _strip_html(raw_text)
        if text:
            labels.append(f"{idx}. {text}")
    return "\n".join(labels)


def _correct_answers(answers: Any) -> str:
    if not isinstance(answers, list):
        return ""
    correct: list[str] = []
    for answer in answers:
        if not isinstance(answer, dict):
            continue
        weight = answer.get("weight")
        try:
            is_correct = float(weight or 0) > 0
        except (TypeError, ValueError):
            is_correct = False
        if is_correct:
            raw_text = (
                answer.get("text")
                or answer.get("html")
                or answer.get("answer_text")
                or answer.get("left")
                or answer.get("right")
                or answer.get("exact")
                or ""
            )
            text = _strip_html(raw_text)
            if text:
                correct.append(text)
    return "; ".join(correct)


def _strip_html(value: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", _safe(value))).strip()


def _sheet_faculty_quiz_images(
    wb,
    helpers: dict[str, Any],
    *,
    questions: list[dict[str, Any]],
    images_by_content_id: dict[str, list[dict[str, Any]]],
) -> None:
    quiz_ws = wb.create_sheet("Quiz & Question Banks")
    quiz_headers = [
        "Quiz / Question Bank", "Question #", "Question Text", "Answer Choices",
        "Correct Answer", "Image", "Alt Text Draft (AI)", "Long Description Draft (AI)",
        "Checked Alt Text", "Checked Long Description (if needed)", "Image Source (do not edit)",
    ]
    _style_header_row(quiz_ws, quiz_headers, helpers)
    row_idx = 2
    for question in questions:
        metadata = _metadata(question)
        question_images = images_by_content_id.get(str(question.get("id") or "")) or [None]
        for image in question_images:
            image_row = image or {}
            _write_cell(quiz_ws, row_idx, 1, metadata.get("parent_quiz_title"), helpers)
            _write_cell(quiz_ws, row_idx, 2, metadata.get("position") or metadata.get("question_id"), helpers)
            _write_cell(quiz_ws, row_idx, 3, _strip_html(metadata.get("question_text")), helpers)
            _write_cell(quiz_ws, row_idx, 4, _format_answers(metadata.get("answers")), helpers)
            _write_cell(quiz_ws, row_idx, 5, _correct_answers(metadata.get("answers")), helpers)
            _write_cell(quiz_ws, row_idx, 6, "", helpers)
            if image:
                _embed_thumbnail(quiz_ws, row=row_idx, col=6, image=image_row, helpers=helpers)
            _write_cell(quiz_ws, row_idx, 7, image_row.get("edited_alt_text") or image_row.get("existing_alt_text"), helpers)
            _write_cell(quiz_ws, row_idx, 8, image_row.get("long_description"), helpers)
            _write_cell(quiz_ws, row_idx, 9, "", helpers)
            _write_cell(quiz_ws, row_idx, 10, "", helpers)
            _write_cell(quiz_ws, row_idx, 11, image_row.get("canvas_url"), helpers)
            row_idx += 1
    _auto_col_widths(quiz_ws, helpers)
    quiz_ws.column_dimensions["F"].width = 20
    quiz_ws.column_dimensions["K"].hidden = True


def _sheet_faculty_content_images(
    wb,
    helpers: dict[str, Any],
    *,
    images: list[dict[str, Any]],
    item_by_id: dict[str, dict[str, Any]],
    file_by_canvas_id: dict[str, dict[str, Any]],
) -> None:
    image_ws = wb.create_sheet("Content Images")
    image_headers = [
        "Page / Content Title", "Content Type", "Image", "Alt Text Draft (AI)",
        "Long Description Draft (AI)", "Checked Alt Text", "Checked Long Description (if needed)",
        "Image Source (do not edit)",
    ]
    _style_header_row(image_ws, image_headers, helpers)
    row_idx = 2
    for image in images:
        item = item_by_id.get(str(image.get("content_item_id") or "")) or {}
        if item.get("content_type") == "quiz_question":
            continue
        _write_cell(image_ws, row_idx, 1, item.get("title") or _image_display_name(image, file_by_canvas_id), helpers)
        _write_cell(image_ws, row_idx, 2, str(item.get("content_type") or "").title(), helpers)
        _write_cell(image_ws, row_idx, 3, "", helpers)
        _embed_thumbnail(image_ws, row=row_idx, col=3, image=image, helpers=helpers)
        _write_cell(image_ws, row_idx, 4, image.get("edited_alt_text") or image.get("existing_alt_text"), helpers)
        _write_cell(image_ws, row_idx, 5, image.get("long_description"), helpers)
        _write_cell(image_ws, row_idx, 6, "", helpers)
        _write_cell(image_ws, row_idx, 7, "", helpers)
        _write_cell(image_ws, row_idx, 8, image.get("canvas_url"), helpers)
        row_idx += 1
    _auto_col_widths(image_ws, helpers)
    image_ws.column_dimensions["C"].width = 20
    image_ws.column_dimensions["H"].hidden = True


def build_content_inventory_workbook(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    helpers = _workbook_helpers()
    wb = helpers["Workbook"]()
    wb.remove(wb.active)
    course = _course_context(supabase, session_id=session_id, user_id=user_id)
    items, decisions = _content_inventory_rows(supabase, session_id=session_id, user_id=user_id)
    item_by_id = {str(row["id"]): row for row in items if row.get("id")}
    file_by_canvas_id = {
        str(row["canvas_id"]): row
        for row in items
        if row.get("content_type") == "file" and row.get("canvas_id")
    }
    images = _fetch_all(
        supabase.table("course_images").select(
            "id, content_item_id, canvas_file_id, canvas_url, existing_alt_text, edited_alt_text, long_description, is_decorative, review_action, is_broken, r2_thumb_key, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).order("created_at")
    )
    findings = _health_findings(supabase, session_id=session_id, user_id=user_id)

    _sheet_summary(wb, helpers, course=course, items=items, images=images, findings=findings)
    _sheet_content_inventory(wb, helpers, items=items, decisions=decisions)
    _sheet_wcag_issues(wb, helpers, findings=findings, item_by_id=item_by_id)
    _sheet_image_alt_text(wb, helpers, images=images, item_by_id=item_by_id, file_by_canvas_id=file_by_canvas_id)
    _sheet_link_issues(wb, helpers, findings=findings, item_by_id=item_by_id)
    _sheet_files(wb, helpers, items=items, decisions=decisions)

    return _xlsx_bytes(wb), XLSX_MEDIA_TYPE, "content_inventory.xlsx"


def build_faculty_review_workbook(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    helpers = _workbook_helpers()
    wb = helpers["Workbook"]()
    wb.remove(wb.active)
    items, decisions = _content_inventory_rows(supabase, session_id=session_id, user_id=user_id)
    item_by_id = {str(row["id"]): row for row in items if row.get("id")}
    file_by_canvas_id = {
        str(row["canvas_id"]): row
        for row in items
        if row.get("content_type") == "file" and row.get("canvas_id")
    }

    images = _fetch_all(
        supabase.table("course_images").select(
            "id, content_item_id, canvas_file_id, canvas_url, existing_alt_text, edited_alt_text, long_description, is_decorative, review_action, is_broken, r2_thumb_key, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).order("created_at")
    )
    questions = _fetch_all(
        supabase.table("course_content_items").select(
            "id, canvas_id, title, metadata, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).eq(
            "content_type", "quiz_question"
        ).order("title")
    )
    images_by_content_id: dict[str, list[dict[str, Any]]] = {}
    for image in images:
        if image.get("content_item_id"):
            images_by_content_id.setdefault(str(image["content_item_id"]), []).append(image)

    _sheet_content_inventory(wb, helpers, items=items, decisions=decisions)
    _sheet_faculty_quiz_images(wb, helpers, questions=questions, images_by_content_id=images_by_content_id)
    _sheet_faculty_content_images(wb, helpers, images=images, item_by_id=item_by_id, file_by_canvas_id=file_by_canvas_id)
    _sheet_files(wb, helpers, items=items, decisions=decisions)

    return _xlsx_bytes(wb), XLSX_MEDIA_TYPE, "faculty_review.xlsx"


def build_content_inventory_csv(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    items = _fetch_all(
        supabase.table("course_content_items").select(
            "id, canvas_id, content_type, title, canvas_url, published, module_name, is_orphaned, duplicate_group_key, metadata, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).order("content_type").order("title")
    )
    decisions = _decision_map(
        supabase,
        session_id=session_id,
        user_id=user_id,
        content_item_ids=[row["id"] for row in items if row.get("id")],
    )
    rows = []
    for item in items:
        metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
        linked_from = metadata.get("linked_from") if isinstance(metadata.get("linked_from"), list) else []
        decision = decisions.get(item.get("id")) or {}
        rows.append({
            "type": item.get("content_type"),
            "title": item.get("title") or "Untitled",
            "status": _content_status(item),
            "module_or_location": item.get("module_name") or ("; ".join(linked_from[:5]) if linked_from else ""),
            "decision": decision.get("action") or "",
            "decision_reason": decision.get("reason") or "",
            "applied_to_canvas": decision.get("applied_to_canvas"),
            "canvas_id": item.get("canvas_id"),
            "canvas_url": item.get("canvas_url"),
            "updated_at": item.get("updated_at"),
        })
    headers = [
        "type", "title", "status", "module_or_location", "decision", "decision_reason",
        "applied_to_canvas", "canvas_id", "canvas_url", "updated_at",
    ]
    return _csv_bytes(rows, headers), "text/csv", "content_inventory.csv"


def build_faculty_review_csv(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    images = _fetch_all(
        supabase.table("course_images").select(
            "id, canvas_file_id, canvas_url, existing_alt_text, edited_alt_text, long_description, is_decorative, review_action, is_broken, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).order("created_at")
    )
    questions = _fetch_all(
        supabase.table("course_content_items").select(
            "id, canvas_id, title, metadata, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).eq(
            "content_type", "quiz_question"
        ).order("title")
    )
    rows = []
    for image in images:
        rows.append({
            "review_type": "image",
            "title": image.get("canvas_file_id") or image.get("canvas_url"),
            "current_text": image.get("existing_alt_text") or "",
            "draft_text": image.get("edited_alt_text") or "",
            "long_description": image.get("long_description") or "",
            "decision": image.get("review_action") or "keep",
            "context": "decorative" if image.get("is_decorative") else "content image",
            "canvas_reference": image.get("canvas_url"),
            "updated_at": image.get("updated_at"),
        })
    for question in questions:
        metadata = question.get("metadata") if isinstance(question.get("metadata"), dict) else {}
        rows.append({
            "review_type": "quiz_question",
            "title": question.get("title") or metadata.get("parent_quiz_title") or "Quiz question",
            "current_text": metadata.get("question_name") or "",
            "draft_text": metadata.get("question_text") or "",
            "long_description": "",
            "decision": "keep",
            "context": metadata.get("parent_quiz_title") or "",
            "canvas_reference": question.get("canvas_id"),
            "updated_at": question.get("updated_at"),
        })
    headers = [
        "review_type", "title", "current_text", "draft_text", "long_description",
        "decision", "context", "canvas_reference", "updated_at",
    ]
    return _csv_bytes(rows, headers), "text/csv", "faculty_review.csv"


def build_latest_transfer_report_json(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    result = supabase.table("background_jobs").select(
        "id, job_type, status, result, error_message, queued_at, started_at, finished_at"
    ).eq("session_id", session_id).eq("user_id", user_id).in_(
        "job_type", ["transfer_target_push", "transfer_same_course_push", "transfer_course_copy"]
    ).order("queued_at", desc=True).limit(1).execute()
    job = result.data[0] if result.data else None
    return _json_bytes({"generated_at": datetime.now(timezone.utc).isoformat(), "job": job}), "application/json", "latest_transfer_report.json"


TRANSFER_REPORT_SECTIONS = [
    ("errors", "Errors"),
    ("warnings", "Warnings"),
    ("protected", "Protected Items"),
    ("skipped", "Skipped Items"),
    ("updated", "Updated"),
    ("created", "Created"),
    ("deleted", "Deleted"),
    ("migrated_files", "Migrated Files"),
    ("placed", "Placed in Modules"),
]


def _latest_transfer_job(supabase, *, session_id: str, user_id: str) -> dict[str, Any] | None:
    result = supabase.table("background_jobs").select(
        "id, job_type, status, result, error_message, queued_at, started_at, finished_at"
    ).eq("session_id", session_id).eq("user_id", user_id).in_(
        "job_type", ["transfer_target_push", "transfer_same_course_push", "transfer_course_copy"]
    ).order("queued_at", desc=True).limit(1).execute()
    return result.data[0] if result.data else None


def _transfer_result(job: dict[str, Any] | None) -> dict[str, Any]:
    result = job.get("result") if isinstance(job, dict) and isinstance(job.get("result"), dict) else {}
    return result if isinstance(result, dict) else {}


def _transfer_report(job: dict[str, Any] | None) -> dict[str, list[dict[str, Any]]]:
    result = _transfer_result(job)
    report = result.get("report") if isinstance(result.get("report"), dict) else {}
    normalized: dict[str, list[dict[str, Any]]] = {}
    for key, value in report.items():
        if isinstance(value, list):
            normalized[str(key)] = [row for row in value if isinstance(row, dict)]
    return normalized


def _transfer_summary(job: dict[str, Any] | None) -> dict[str, Any]:
    result = _transfer_result(job)
    summary = result.get("summary") if isinstance(result.get("summary"), dict) else {}
    return summary if isinstance(summary, dict) else {}


def _sheet_transfer_summary(wb, helpers: dict[str, Any], *, job: dict[str, Any] | None, report: dict[str, list[dict[str, Any]]]) -> None:
    ws = wb.create_sheet("Summary", 0)
    result = _transfer_result(job)
    summary = _transfer_summary(job)
    target_course = result.get("target_course") if isinstance(result.get("target_course"), dict) else {}
    rows = [
        ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Job ID", job.get("id") if job else ""),
        ("Job Type", str(job.get("job_type") or "").replace("_", " ").title() if job else "No transfer job found"),
        ("Job Status", job.get("status") if job else ""),
        ("Result Status", result.get("status") or ""),
        ("Queued At", job.get("queued_at") if job else ""),
        ("Started At", job.get("started_at") if job else ""),
        ("Finished At", job.get("finished_at") if job else ""),
        ("Target Course", target_course.get("name") or ""),
        ("Target Canvas Course ID", target_course.get("canvas_course_id") or summary.get("target_canvas_course_id") or ""),
        ("Error Message", job.get("error_message") if job else ""),
        ("", ""),
        ("REPORT COUNTS", ""),
    ]
    for key, label in TRANSFER_REPORT_SECTIONS:
        rows.append((label, len(report.get(key) or [])))
    extra_keys = sorted(key for key in report.keys() if key not in {section[0] for section in TRANSFER_REPORT_SECTIONS})
    for key in extra_keys:
        rows.append((key.replace("_", " ").title(), len(report.get(key) or [])))

    rows.append(("", ""))
    rows.append(("JOB SUMMARY", ""))
    for key, value in sorted(summary.items()):
        if isinstance(value, (dict, list)):
            value = json.dumps(value, default=str)
        rows.append((str(key).replace("_", " ").title(), value))

    for row_idx, (label, value) in enumerate(rows, 1):
        _write_cell(ws, row_idx, 1, label, helpers)
        _write_cell(ws, row_idx, 2, value, helpers)
        if label and label == label.upper():
            ws.cell(row=row_idx, column=1).font = helpers["section_font"]
    _auto_col_widths(ws, helpers)


def _safe_sheet_title(value: str) -> str:
    title = re.sub(r"[\[\]*?:/\\]", " ", value).strip() or "Report"
    return title[:31]


def _sheet_transfer_report_section(
    wb,
    helpers: dict[str, Any],
    *,
    title: str,
    rows: list[dict[str, Any]],
    fill_key: str | None = None,
) -> None:
    ws = wb.create_sheet(_safe_sheet_title(title))
    headers = ["Title", "Content Type", "Action", "Status", "Reason / Details", "Canvas Link"]
    _style_header_row(ws, headers, helpers)
    for row_idx, item in enumerate(rows, 2):
        canvas_url = item.get("canvas_url")
        status = str(item.get("status") or "")
        _write_cell(ws, row_idx, 1, item.get("title") or "Untitled item", helpers)
        _write_cell(ws, row_idx, 2, item.get("content_type") or "", helpers)
        _write_cell(ws, row_idx, 3, str(item.get("action") or "").replace("_", " ").title(), helpers)
        _write_cell(ws, row_idx, 4, status.replace("_", " ").title(), helpers, fill_key=status or fill_key)
        _write_cell(ws, row_idx, 5, item.get("reason") or "", helpers)
        _write_cell(ws, row_idx, 6, "Open in Canvas" if canvas_url else "", helpers, hyperlink=canvas_url)
    _auto_col_widths(ws, helpers)


def build_latest_transfer_report_workbook(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    helpers = _workbook_helpers()
    wb = helpers["Workbook"]()
    wb.remove(wb.active)
    job = _latest_transfer_job(supabase, session_id=session_id, user_id=user_id)
    report = _transfer_report(job)
    _sheet_transfer_summary(wb, helpers, job=job, report=report)

    section_fill_keys = {
        "errors": "critical",
        "warnings": "warning",
        "protected": "warning",
        "skipped": "warning",
        "updated": "linked",
        "created": "in_module",
        "deleted": "orphaned",
        "migrated_files": "in_module",
        "placed": "linked",
    }
    rendered_keys: set[str] = set()
    for key, title in TRANSFER_REPORT_SECTIONS:
        rows = report.get(key) or []
        if rows:
            _sheet_transfer_report_section(
                wb,
                helpers,
                title=title,
                rows=rows,
                fill_key=section_fill_keys.get(key),
            )
            rendered_keys.add(key)
    for key, rows in sorted(report.items()):
        if key in rendered_keys or not rows:
            continue
        _sheet_transfer_report_section(
            wb,
            helpers,
            title=key.replace("_", " ").title(),
            rows=rows,
        )

    return _xlsx_bytes(wb), XLSX_MEDIA_TYPE, "latest_transfer_report.xlsx"


def build_health_summary_csv(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    run_result = supabase.table("health_runs").select(
        "id, status, items_scanned, duration_ms, summary, created_at, finished_at"
    ).eq("session_id", session_id).eq("user_id", user_id).order("created_at", desc=True).limit(1).execute()
    run = run_result.data[0] if run_result.data else None
    if not run:
        return _csv_bytes([], ["finding_type", "finding_code", "severity", "description", "content_item_id"]), "text/csv", "health_summary.csv"
    findings = _fetch_all(
        supabase.table("health_findings").select(
            "finding_type, finding_code, severity, description, content_item_id, created_at"
        ).eq("session_id", session_id).eq("health_run_id", run["id"]).order("severity")
    )
    headers = ["finding_type", "finding_code", "severity", "description", "content_item_id", "created_at"]
    return _csv_bytes(findings, headers), "text/csv", "health_summary.csv"


def build_health_summary_workbook(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    helpers = _workbook_helpers()
    wb = helpers["Workbook"]()
    wb.remove(wb.active)
    course = _course_context(supabase, session_id=session_id, user_id=user_id)
    run_result = supabase.table("health_runs").select(
        "id, status, items_scanned, duration_ms, summary, created_at, finished_at"
    ).eq("session_id", session_id).eq("user_id", user_id).order("created_at", desc=True).limit(1).execute()
    run = run_result.data[0] if run_result.data else None
    items, _decisions = _content_inventory_rows(supabase, session_id=session_id, user_id=user_id)
    item_by_id = {str(row["id"]): row for row in items if row.get("id")}
    file_items = [row for row in items if row.get("content_type") == "file"]
    file_by_canvas_id = {
        str(row["canvas_id"]): row
        for row in file_items
        if row.get("canvas_id")
    }
    images = _fetch_all(
        supabase.table("course_images").select(
            "id, content_item_id, canvas_file_id, canvas_url, existing_alt_text, edited_alt_text, long_description, is_decorative, review_action, is_broken, r2_thumb_key, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).order("created_at")
    )
    findings = _health_findings(supabase, session_id=session_id, user_id=user_id)
    documents = _fetch_all(
        supabase.table("documents").select(
            "id, filename, status, page_count, tag_data, created_at, updated_at"
        ).eq("session_id", session_id).eq("user_id", user_id).order("updated_at", desc=True)
    )

    _sheet_health_summary(wb, helpers, course=course, run=run, items=items, images=images, findings=findings)
    _sheet_wcag_issues(wb, helpers, findings=findings, item_by_id=item_by_id)
    wb["WCAG Issues"].title = "WCAG Findings"
    _sheet_health_image_issues(wb, helpers, images=images, item_by_id=item_by_id, file_by_canvas_id=file_by_canvas_id)
    _sheet_link_issues(wb, helpers, findings=findings, item_by_id=item_by_id)
    _sheet_health_inventory_findings(wb, helpers, findings=findings, item_by_id=item_by_id)
    _sheet_health_files_documents(wb, helpers, file_items=file_items, documents=documents)

    return _xlsx_bytes(wb), XLSX_MEDIA_TYPE, "health_summary.xlsx"


def build_edit_history_csv(supabase, *, session_id: str, user_id: str) -> tuple[bytes, str, str]:
    items = _fetch_all(
        supabase.table("course_content_items").select(
            "id, title, content_type"
        ).eq("session_id", session_id).eq("user_id", user_id)
    )
    item_by_id = {row["id"]: row for row in items if row.get("id")}
    revisions = _fetch_all(
        supabase.table("content_revisions").select(
            "content_item_id, revision_number, change_summary, created_at"
        ).eq("session_id", session_id).eq("user_id", user_id).order("created_at", desc=True)
    )
    rows = []
    for revision in revisions:
        item = item_by_id.get(revision.get("content_item_id")) or {}
        rows.append({
            "content_type": item.get("content_type") or "",
            "title": item.get("title") or "Untitled",
            "revision_number": revision.get("revision_number"),
            "change_summary": revision.get("change_summary") or "",
            "created_at": revision.get("created_at"),
        })
    headers = ["content_type", "title", "revision_number", "change_summary", "created_at"]
    return _csv_bytes(rows, headers), "text/csv", "edit_history.csv"


EXPORT_BUILDERS = {
    "content_inventory": build_content_inventory_workbook,
    "faculty_review": build_faculty_review_workbook,
    "transfer_report": build_latest_transfer_report_workbook,
    "health_summary": build_health_summary_workbook,
    "edit_history": build_edit_history_csv,
}
