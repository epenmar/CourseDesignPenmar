"""Apply Faculty Review workbook edits back to session review tables."""

from __future__ import annotations

import io
from collections.abc import Iterable
from datetime import datetime, timezone
from typing import Any, Literal

from services.inventory_decision_sync import (
    sync_file_decisions_to_image_reviews,
    sync_image_reviews_to_file_decisions,
)


DecisionAction = Literal["keep", "delete", "defer"]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_decision(value: Any) -> DecisionAction | None:
    text = _cell_text(value).casefold()
    if text in {"keep", "kept"}:
        return "keep"
    if text in {"remove", "delete", "deleted"}:
        return "delete"
    if text in {"defer", "deferred"}:
        return "defer"
    return None


def _iter_rows(ws, *, min_row: int) -> Iterable[tuple[int, tuple[Any, ...]]]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True), min_row):
        yield row_idx, row


def _value(row: tuple[Any, ...], index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _parse_image_sheet(
    wb,
    *,
    sheet_name: str,
    start_row: int,
    alt_col: int,
    long_desc_col: int,
    source_col: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    if ws is None:
        return [], [{"sheet": sheet_name, "row": None, "reason": "Worksheet not found"}]

    updates: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    for row_idx, row in _iter_rows(ws, min_row=start_row):
        canvas_url = _cell_text(_value(row, source_col))
        checked_alt = _cell_text(_value(row, alt_col))
        checked_long_desc = _cell_text(_value(row, long_desc_col))
        if not canvas_url and not checked_alt and not checked_long_desc:
            continue
        if not canvas_url:
            skipped.append({"sheet": sheet_name, "row": row_idx, "reason": "Missing hidden image source"})
            continue
        if not checked_alt and not checked_long_desc:
            continue
        updates.append({
            "sheet": sheet_name,
            "row": row_idx,
            "canvas_url": canvas_url,
            "checked_alt_text": checked_alt or None,
            "checked_long_description": checked_long_desc or None,
        })
    return updates, skipped


def _parse_decision_sheet(
    wb,
    *,
    sheet_name: str,
    start_row: int,
    decision_col: int,
    key_col: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    if ws is None:
        return [], [{"sheet": sheet_name, "row": None, "reason": "Worksheet not found"}]

    updates: dict[str, dict[str, Any]] = {}
    skipped: list[dict[str, Any]] = []
    for row_idx, row in _iter_rows(ws, min_row=start_row):
        item_key = _cell_text(_value(row, key_col))
        decision = _normalize_decision(_value(row, decision_col))
        if not item_key and not decision:
            continue
        if not item_key:
            skipped.append({"sheet": sheet_name, "row": row_idx, "reason": "Missing hidden item key"})
            continue
        if decision is None:
            continue
        updates[item_key] = {
            "sheet": sheet_name,
            "row": row_idx,
            "item_key": item_key,
            "action": decision,
        }
    return list(updates.values()), skipped


def parse_faculty_review_workbook(file_bytes: bytes) -> dict[str, Any]:
    """Parse the v2 Faculty Review workbook into image and inventory updates."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("openpyxl is required to upload Faculty Review workbooks") from exc

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    skipped: list[dict[str, Any]] = []

    quiz_image_updates, quiz_skipped = _parse_image_sheet(
        wb,
        sheet_name="Quiz & Question Banks",
        start_row=2,
        alt_col=8,
        long_desc_col=9,
        source_col=10,
    )
    content_image_updates, content_skipped = _parse_image_sheet(
        wb,
        sheet_name="Content Images",
        start_row=2,
        alt_col=5,
        long_desc_col=6,
        source_col=7,
    )
    inventory_decisions, inventory_skipped = _parse_decision_sheet(
        wb,
        sheet_name="Content Inventory",
        start_row=3,
        decision_col=4,
        key_col=9,
    )
    file_decisions, file_skipped = _parse_decision_sheet(
        wb,
        sheet_name="Files",
        start_row=2,
        decision_col=3,
        key_col=7,
    )
    skipped.extend(quiz_skipped)
    skipped.extend(content_skipped)
    skipped.extend(inventory_skipped)
    skipped.extend(file_skipped)

    return {
        "image_updates": quiz_image_updates + content_image_updates,
        "decision_updates": inventory_decisions + file_decisions,
        "skipped": skipped,
    }


def _fetch_all(query, *, page_size: int = 1000) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    offset = 0
    while True:
        result = query.range(offset, offset + page_size - 1).execute()
        batch = result.data or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return rows


def _chunks(values: list[str], size: int = 50) -> Iterable[list[str]]:
    for index in range(0, len(values), size):
        yield values[index:index + size]


def _item_key(row: dict[str, Any]) -> str:
    return f"{row.get('content_type') or ''}-{row.get('canvas_id') or ''}"


def _upsert_decisions(
    supabase,
    *,
    session_id: str,
    user_id: str,
    desired: list[dict[str, str]],
) -> dict[str, Any]:
    if not desired:
        return {"created": 0, "updated": 0}

    deduped: dict[str, dict[str, str]] = {}
    for row in desired:
        if row.get("content_item_id"):
            deduped[row["content_item_id"]] = row
    desired = list(deduped.values())
    content_item_ids = list(dict.fromkeys([row["content_item_id"] for row in desired if row.get("content_item_id")]))
    existing_by_item_id: dict[str, dict[str, Any]] = {}
    for chunk in _chunks(content_item_ids):
        existing = supabase.table("content_inventory_decisions").select(
            "id, content_item_id"
        ).eq("session_id", session_id).eq("user_id", user_id).in_(
            "content_item_id", chunk
        ).execute()
        for row in existing.data or []:
            if row.get("id") and row.get("content_item_id"):
                existing_by_item_id[row["content_item_id"]] = row

    now = _now_iso()
    update_groups: dict[tuple[str, str], list[str]] = {}
    inserts: list[dict[str, Any]] = []
    for row in desired:
        existing = existing_by_item_id.get(row["content_item_id"])
        if existing:
            update_groups.setdefault((row["action"], row["reason"]), []).append(existing["id"])
        else:
            inserts.append({
                "content_item_id": row["content_item_id"],
                "session_id": session_id,
                "user_id": user_id,
                "action": row["action"],
                "reason": row["reason"],
                "updated_at": now,
            })

    updated = 0
    for (action, reason), decision_ids in update_groups.items():
        for chunk in _chunks(decision_ids):
            result = supabase.table("content_inventory_decisions").update({
                "action": action,
                "reason": reason,
                "updated_at": now,
            }).in_("id", chunk).execute()
            updated += len(result.data or chunk)

    created = 0
    for index in range(0, len(inserts), 100):
        result = supabase.table("content_inventory_decisions").insert(inserts[index:index + 100]).execute()
        created += len(result.data or inserts[index:index + 100])

    return {"created": created, "updated": updated}


def apply_faculty_review_workbook(
    supabase,
    *,
    session_id: str,
    user_id: str,
    file_bytes: bytes,
) -> dict[str, Any]:
    parsed = parse_faculty_review_workbook(file_bytes)
    skipped: list[dict[str, Any]] = list(parsed["skipped"])

    image_rows = _fetch_all(
        supabase.table("course_images").select(
            "id, canvas_url, canvas_file_id, content_item_id"
        ).eq("session_id", session_id).eq("user_id", user_id)
    )
    image_by_url = {
        str(row["canvas_url"]): row
        for row in image_rows
        if row.get("id") and row.get("canvas_url")
    }

    image_update_count = 0
    image_file_ids: list[str] = []
    now = _now_iso()
    for update in parsed["image_updates"]:
        image = image_by_url.get(update["canvas_url"])
        if not image:
            skipped.append({
                "sheet": update["sheet"],
                "row": update["row"],
                "reason": "Image source did not match this session",
            })
            continue
        payload: dict[str, Any] = {"review_action": "keep", "updated_at": now}
        if update.get("checked_alt_text") is not None:
            payload["edited_alt_text"] = update["checked_alt_text"][:1000]
        if update.get("checked_long_description") is not None:
            payload["long_description"] = update["checked_long_description"][:8000]
        supabase.table("course_images").update(payload).eq("id", image["id"]).execute()
        image_update_count += 1
        if image.get("canvas_file_id"):
            image_file_ids.append(str(image["canvas_file_id"]))

    item_rows = _fetch_all(
        supabase.table("course_content_items").select(
            "id, canvas_id, content_type"
        ).eq("session_id", session_id).eq("user_id", user_id)
    )
    item_by_key = {
        _item_key(row): row
        for row in item_rows
        if row.get("id") and row.get("canvas_id") and row.get("content_type")
    }

    desired_decisions: list[dict[str, str]] = []
    file_decision_item_ids: list[str] = []
    for decision in parsed["decision_updates"]:
        item = item_by_key.get(decision["item_key"])
        if not item:
            skipped.append({
                "sheet": decision["sheet"],
                "row": decision["row"],
                "reason": "Item key did not match this session",
            })
            continue
        desired_decisions.append({
            "content_item_id": item["id"],
            "action": decision["action"],
            "reason": "Applied from Faculty Review workbook upload",
        })
        if item.get("content_type") == "file":
            file_decision_item_ids.append(item["id"])

    deduped_decisions = {
        row["content_item_id"]: row
        for row in desired_decisions
    }
    desired_decisions = list(deduped_decisions.values())
    file_decision_item_ids = [
        item_id
        for item_id in dict.fromkeys(file_decision_item_ids)
        if item_id in deduped_decisions
    ]

    decision_result = _upsert_decisions(
        supabase,
        session_id=session_id,
        user_id=user_id,
        desired=desired_decisions,
    )

    file_sync_result = {"updated_image_count": 0, "canvas_file_ids": []}

    # Sync file decisions by action so image inventory rows stay aligned with
    # uploaded Files-tab changes without forcing every file to the same state.
    for action in ("keep", "delete", "defer"):
        action_file_ids = [
            row["content_item_id"]
            for row in desired_decisions
            if row["content_item_id"] in file_decision_item_ids and row["action"] == action
        ]
        if action_file_ids:
            next_sync = sync_file_decisions_to_image_reviews(
                supabase,
                session_id=session_id,
                user_id=user_id,
                content_item_ids=action_file_ids,
                action=action,  # type: ignore[arg-type]
            )
            file_sync_result["updated_image_count"] += next_sync.get("updated_image_count", 0)
            file_sync_result["canvas_file_ids"].extend(next_sync.get("canvas_file_ids", []))

    image_sync_result = sync_image_reviews_to_file_decisions(
        supabase,
        session_id=session_id,
        user_id=user_id,
        canvas_file_ids=image_file_ids,
    )

    supabase.table("platform_events").insert({
        "user_id": user_id,
        "session_id": session_id,
        "event_type": "faculty_review_upload_applied",
        "properties": {
            "image_updates": image_update_count,
            "decision_updates": len(desired_decisions),
            "decision_created": decision_result["created"],
            "decision_updated": decision_result["updated"],
            "skipped": len(skipped),
        },
    }).execute()

    return {
        "image_updates": image_update_count,
        "decision_updates": len(desired_decisions),
        "decision_created": decision_result["created"],
        "decision_updated": decision_result["updated"],
        "synced_image_reviews": file_sync_result.get("updated_image_count", 0),
        "synced_file_decisions": image_sync_result.get("updated_file_decision_count", 0),
        "skipped": skipped[:25],
        "skipped_count": len(skipped),
    }
